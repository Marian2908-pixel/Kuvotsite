from fastapi import FastAPI, APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import httpx

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Nova Poshta API configuration
NOVA_POSHTA_API_KEY = os.environ.get('NOVA_POSHTA_API_KEY', '')
NOVA_POSHTA_API_URL = "https://api.novaposhta.ua/v2.0/json/"

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ========== MODELS ==========

class PriceItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    size: str
    cost_price: float
    sell_price: float
    lacquer_cost: float
    lacquer_price: float
    packaging_cost: float
    packaging_price: float
    frame_1_10_cost: float
    frame_1_10_price: float
    frame_11_14_cost: float
    frame_11_14_price: float

class PriceItemCreate(BaseModel):
    size: str
    cost_price: float
    sell_price: float
    lacquer_cost: float = 0
    lacquer_price: float = 0
    packaging_cost: float = 0
    packaging_price: float = 0
    frame_1_10_cost: float = 0
    frame_1_10_price: float = 0
    frame_11_14_cost: float = 0
    frame_11_14_price: float = 0

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    cost_price: float
    sell_price: float
    category: str = "інше"

class ProductCreate(BaseModel):
    name: str
    cost_price: float
    sell_price: float
    category: str = "інше"

class OrderItem(BaseModel):
    size: Optional[str] = None
    product_id: Optional[str] = None
    product_name: Optional[str] = None
    quantity: int = 1
    unit_price: float
    unit_cost: float
    with_lacquer: bool = False
    with_packaging: bool = False
    frame_type: Optional[str] = None  # "1-10", "11-14", or None
    lacquer_price: float = 0
    lacquer_cost: float = 0
    packaging_price: float = 0
    packaging_cost: float = 0
    frame_price: float = 0
    frame_cost: float = 0
    total_price: float = 0
    total_cost: float = 0
    profit: float = 0

class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_date: str
    month: str
    painting_name: str
    order_type: str  # цифрова, друк, оригінал
    items: List[OrderItem]
    total_amount: float  # основна сума замовлення
    total_cost: float  # собівартість
    profit: float  # чистий прибуток
    sales_channel: str  # Instagram, Messenger, Viber/Telegram
    status: str  # нове, оплачено, виконано, скасовано
    comment: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    # New fields for discount and extra income
    extra_income: float = 0  # додатковий дохід
    discounted_amount: Optional[float] = None  # сума зі знижкою
    discount: float = 0  # розмір знижки
    net_income: float = 0  # чистий дохід (final amount for analytics)

class OrderCreate(BaseModel):
    order_date: str
    month: Optional[str] = None
    painting_name: str
    order_type: str
    items: List[OrderItem]
    sales_channel: str
    status: str = "нове"
    comment: Optional[str] = None
    extra_income: float = 0
    discounted_amount: Optional[float] = None

class OrderUpdate(BaseModel):
    order_date: Optional[str] = None
    month: Optional[str] = None
    painting_name: Optional[str] = None
    order_type: Optional[str] = None
    items: Optional[List[OrderItem]] = None
    sales_channel: Optional[str] = None
    status: Optional[str] = None
    comment: Optional[str] = None
    extra_income: Optional[float] = None
    discounted_amount: Optional[float] = None

# ========== PRICE CATALOG ==========

@api_router.get("/prices", response_model=List[PriceItem])
async def get_prices():
    prices = await db.price_catalog.find({}, {"_id": 0}).to_list(100)
    return prices

@api_router.post("/prices", response_model=PriceItem)
async def create_price(data: PriceItemCreate):
    price_obj = PriceItem(**data.model_dump())
    doc = price_obj.model_dump()
    await db.price_catalog.insert_one(doc)
    return price_obj

@api_router.put("/prices/{price_id}", response_model=PriceItem)
async def update_price(price_id: str, data: PriceItemCreate):
    result = await db.price_catalog.find_one_and_update(
        {"id": price_id},
        {"$set": data.model_dump()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Ціну не знайдено")
    result.pop("_id", None)
    return result

@api_router.delete("/prices/{price_id}")
async def delete_price(price_id: str):
    result = await db.price_catalog.delete_one({"id": price_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ціну не знайдено")
    return {"message": "Видалено"}

@api_router.post("/prices/seed")
async def seed_prices():
    """Seed initial price catalog"""
    existing = await db.price_catalog.count_documents({})
    if existing > 0:
        return {"message": "Каталог вже заповнено", "count": existing}
    
    prices_data = [
        {"size": "20х30", "cost_price": 170, "sell_price": 420, "lacquer_cost": 30, "lacquer_price": 80, "packaging_cost": 30, "packaging_price": 90, "frame_1_10_cost": 180, "frame_1_10_price": 360, "frame_11_14_cost": 130, "frame_11_14_price": 260},
        {"size": "25х30", "cost_price": 180, "sell_price": 520, "lacquer_cost": 30, "lacquer_price": 80, "packaging_cost": 30, "packaging_price": 90, "frame_1_10_cost": 200, "frame_1_10_price": 390, "frame_11_14_cost": 145, "frame_11_14_price": 290},
        {"size": "30х35", "cost_price": 220, "sell_price": 580, "lacquer_cost": 30, "lacquer_price": 100, "packaging_cost": 30, "packaging_price": 110, "frame_1_10_cost": 235, "frame_1_10_price": 440, "frame_11_14_cost": 170, "frame_11_14_price": 340},
        {"size": "30х40", "cost_price": 240, "sell_price": 650, "lacquer_cost": 30, "lacquer_price": 100, "packaging_cost": 30, "packaging_price": 110, "frame_1_10_cost": 250, "frame_1_10_price": 490, "frame_11_14_cost": 180, "frame_11_14_price": 380},
        {"size": "30х45", "cost_price": 250, "sell_price": 750, "lacquer_cost": 30, "lacquer_price": 100, "packaging_cost": 30, "packaging_price": 120, "frame_1_10_cost": 270, "frame_1_10_price": 540, "frame_11_14_cost": 195, "frame_11_14_price": 400},
        {"size": "35х45", "cost_price": 260, "sell_price": 850, "lacquer_cost": 30, "lacquer_price": 100, "packaging_cost": 35, "packaging_price": 120, "frame_1_10_cost": 290, "frame_1_10_price": 580, "frame_11_14_cost": 210, "frame_11_14_price": 420},
        {"size": "35х50", "cost_price": 275, "sell_price": 950, "lacquer_cost": 30, "lacquer_price": 100, "packaging_cost": 35, "packaging_price": 120, "frame_1_10_cost": 305, "frame_1_10_price": 610, "frame_11_14_cost": 220, "frame_11_14_price": 440},
        {"size": "40х50", "cost_price": 290, "sell_price": 1050, "lacquer_cost": 35, "lacquer_price": 120, "packaging_cost": 35, "packaging_price": 130, "frame_1_10_cost": 325, "frame_1_10_price": 650, "frame_11_14_cost": 235, "frame_11_14_price": 470},
        {"size": "40х60", "cost_price": 340, "sell_price": 1180, "lacquer_cost": 40, "lacquer_price": 120, "packaging_cost": 35, "packaging_price": 130, "frame_1_10_cost": 360, "frame_1_10_price": 720, "frame_11_14_cost": 260, "frame_11_14_price": 520},
        {"size": "50х60", "cost_price": 370, "sell_price": 1280, "lacquer_cost": 45, "lacquer_price": 130, "packaging_cost": 50, "packaging_price": 140, "frame_1_10_cost": 395, "frame_1_10_price": 790, "frame_11_14_cost": 285, "frame_11_14_price": 570},
        {"size": "50х70", "cost_price": 425, "sell_price": 1380, "lacquer_cost": 50, "lacquer_price": 130, "packaging_cost": 60, "packaging_price": 140, "frame_1_10_cost": 430, "frame_1_10_price": 860, "frame_11_14_cost": 310, "frame_11_14_price": 620},
        {"size": "50х75", "cost_price": 455, "sell_price": 1480, "lacquer_cost": 50, "lacquer_price": 130, "packaging_cost": 60, "packaging_price": 140, "frame_1_10_cost": 450, "frame_1_10_price": 900, "frame_11_14_cost": 325, "frame_11_14_price": 650},
        {"size": "60х70", "cost_price": 510, "sell_price": 1600, "lacquer_cost": 65, "lacquer_price": 140, "packaging_cost": 60, "packaging_price": 150, "frame_1_10_cost": 470, "frame_1_10_price": 950, "frame_11_14_cost": 340, "frame_11_14_price": 700},
        {"size": "60х80", "cost_price": 575, "sell_price": 1750, "lacquer_cost": 70, "lacquer_price": 140, "packaging_cost": 65, "packaging_price": 150, "frame_1_10_cost": 505, "frame_1_10_price": 1010, "frame_11_14_cost": 365, "frame_11_14_price": 730},
        {"size": "60х90", "cost_price": 640, "sell_price": 1890, "lacquer_cost": 75, "lacquer_price": 140, "packaging_cost": 65, "packaging_price": 160, "frame_1_10_cost": 540, "frame_1_10_price": 1080, "frame_11_14_cost": 390, "frame_11_14_price": 750},
        {"size": "70х80", "cost_price": 675, "sell_price": 1890, "lacquer_cost": 80, "lacquer_price": 140, "packaging_cost": 120, "packaging_price": 160, "frame_1_10_cost": 540, "frame_1_10_price": 1080, "frame_11_14_cost": 390, "frame_11_14_price": 750},
        {"size": "70х90", "cost_price": 755, "sell_price": 1980, "lacquer_cost": 85, "lacquer_price": 150, "packaging_cost": 120, "packaging_price": 170, "frame_1_10_cost": 575, "frame_1_10_price": 1150, "frame_11_14_cost": 415, "frame_11_14_price": 780},
        {"size": "70х100", "cost_price": 835, "sell_price": 2050, "lacquer_cost": 90, "lacquer_price": 180, "packaging_cost": 135, "packaging_price": 210, "frame_1_10_cost": 610, "frame_1_10_price": 1220, "frame_11_14_cost": 440, "frame_11_14_price": 830},
        {"size": "80х100", "cost_price": 960, "sell_price": 2190, "lacquer_cost": 105, "lacquer_price": 190, "packaging_cost": 135, "packaging_price": 220, "frame_1_10_cost": 650, "frame_1_10_price": 1300, "frame_11_14_cost": 470, "frame_11_14_price": 940},
        {"size": "80х120", "cost_price": 1145, "sell_price": 2490, "lacquer_cost": 120, "lacquer_price": 230, "packaging_cost": 140, "packaging_price": 260, "frame_1_10_cost": 720, "frame_1_10_price": 1440, "frame_11_14_cost": 520, "frame_11_14_price": 1040},
        {"size": "90х120", "cost_price": 1280, "sell_price": 2690, "lacquer_cost": 140, "lacquer_price": 250, "packaging_cost": 155, "packaging_price": 280, "frame_1_10_cost": 755, "frame_1_10_price": 1510, "frame_11_14_cost": 545, "frame_11_14_price": 1090},
        {"size": "100х150", "cost_price": 1785, "sell_price": 3240, "lacquer_cost": 185, "lacquer_price": 340, "packaging_cost": 265, "packaging_price": 390, "frame_1_10_cost": 900, "frame_1_10_price": 1800, "frame_11_14_cost": 650, "frame_11_14_price": 1300},
    ]
    
    for data in prices_data:
        price_obj = PriceItem(**data)
        await db.price_catalog.insert_one(price_obj.model_dump())
    
    return {"message": "Каталог цін заповнено", "count": len(prices_data)}

# ========== PRODUCTS ==========

@api_router.get("/products", response_model=List[Product])
async def get_products():
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    return products

@api_router.post("/products", response_model=Product)
async def create_product(data: ProductCreate):
    product_obj = Product(**data.model_dump())
    doc = product_obj.model_dump()
    await db.products.insert_one(doc)
    return product_obj

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, data: ProductCreate):
    result = await db.products.find_one_and_update(
        {"id": product_id},
        {"$set": data.model_dump()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Товар не знайдено")
    result.pop("_id", None)
    return result

@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str):
    result = await db.products.delete_one({"id": product_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Товар не знайдено")
    return {"message": "Видалено"}

# ========== ORDERS ==========

def calculate_order_totals(items: List[OrderItem]) -> tuple:
    total_amount = 0
    total_cost = 0
    for item in items:
        item_total_price = item.unit_price * item.quantity
        item_total_cost = item.unit_cost * item.quantity
        
        if item.with_lacquer:
            item_total_price += item.lacquer_price * item.quantity
            item_total_cost += item.lacquer_cost * item.quantity
        if item.with_packaging:
            item_total_price += item.packaging_price * item.quantity
            item_total_cost += item.packaging_cost * item.quantity
        if item.frame_type:
            item_total_price += item.frame_price * item.quantity
            item_total_cost += item.frame_cost * item.quantity
        
        item.total_price = item_total_price
        item.total_cost = item_total_cost
        item.profit = item_total_price - item_total_cost
        
        total_amount += item_total_price
        total_cost += item_total_cost
    
    return total_amount, total_cost, total_amount - total_cost

def extract_month(order_date: str) -> str:
    try:
        dt = datetime.fromisoformat(order_date.replace('Z', '+00:00'))
        months_ua = ["Січень", "Лютий", "Березень", "Квітень", "Травень", "Червень",
                     "Липень", "Серпень", "Вересень", "Жовтень", "Листопад", "Грудень"]
        return f"{months_ua[dt.month - 1]} {dt.year}"
    except:
        return ""

@api_router.get("/orders", response_model=List[Order])
async def get_orders(
    month: Optional[str] = None,
    order_type: Optional[str] = None,
    size: Optional[str] = None,
    status: Optional[str] = None,
    sales_channel: Optional[str] = None
):
    query = {}
    if month:
        query["month"] = month
    if order_type:
        query["order_type"] = order_type
    if status:
        query["status"] = status
    if sales_channel:
        query["sales_channel"] = sales_channel
    if size:
        query["items.size"] = size
    
    orders = await db.orders.find(query, {"_id": 0}).sort("order_date", -1).to_list(1000)
    return orders

@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Замовлення не знайдено")
    return order

@api_router.post("/orders", response_model=Order)
async def create_order(data: OrderCreate):
    order_dict = data.model_dump()
    
    # Calculate totals
    items = [OrderItem(**item) if isinstance(item, dict) else item for item in order_dict["items"]]
    total_amount, total_cost, profit = calculate_order_totals(items)
    
    # Extract month
    month = order_dict.get("month") or extract_month(order_dict["order_date"])
    
    # Calculate discount and net income
    extra_income = order_dict.get("extra_income", 0) or 0
    discounted_amount = order_dict.get("discounted_amount")
    
    # If discounted_amount is provided, use it as the final amount
    if discounted_amount is not None and discounted_amount > 0:
        discount = total_amount - discounted_amount
        final_amount = discounted_amount
    else:
        discount = 0
        final_amount = total_amount
        discounted_amount = None
    
    # Net income = final amount + extra income - cost
    net_income = final_amount + extra_income - total_cost
    
    # Update order_dict with calculated values
    order_dict.update({
        "month": month,
        "items": items,
        "total_amount": total_amount,
        "total_cost": total_cost,
        "profit": profit,
        "extra_income": extra_income,
        "discounted_amount": discounted_amount,
        "discount": discount,
        "net_income": net_income
    })
    
    order_obj = Order(**order_dict)
    
    doc = order_obj.model_dump()
    await db.orders.insert_one(doc)
    return order_obj

@api_router.put("/orders/{order_id}", response_model=Order)
async def update_order(order_id: str, data: OrderUpdate):
    existing = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Замовлення не знайдено")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    
    # Merge with existing data for recalculation
    merged = {**existing, **update_data}
    
    if "items" in update_data:
        items = [OrderItem(**item) if isinstance(item, dict) else item for item in update_data["items"]]
        total_amount, total_cost, profit = calculate_order_totals(items)
        update_data["items"] = [item.model_dump() for item in items]
        update_data["total_amount"] = total_amount
        update_data["total_cost"] = total_cost
        update_data["profit"] = profit
    else:
        total_amount = merged.get("total_amount", 0)
        total_cost = merged.get("total_cost", 0)
        profit = merged.get("profit", 0)
    
    if "order_date" in update_data:
        update_data["month"] = update_data.get("month") or extract_month(update_data["order_date"])
    
    # Recalculate discount and net income
    extra_income = update_data.get("extra_income", merged.get("extra_income", 0)) or 0
    discounted_amount = update_data.get("discounted_amount", merged.get("discounted_amount"))
    
    if discounted_amount is not None and discounted_amount > 0:
        discount = total_amount - discounted_amount
        final_amount = discounted_amount
    else:
        discount = 0
        final_amount = total_amount
        discounted_amount = None
    
    net_income = final_amount + extra_income - total_cost
    
    update_data["extra_income"] = extra_income
    update_data["discounted_amount"] = discounted_amount
    update_data["discount"] = discount
    update_data["net_income"] = net_income
    
    result = await db.orders.find_one_and_update(
        {"id": order_id},
        {"$set": update_data},
        return_document=True
    )
    result.pop("_id", None)
    return result

@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str):
    result = await db.orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Замовлення не знайдено")
    return {"message": "Видалено"}

# ========== ANALYTICS ==========

@api_router.get("/analytics/summary")
async def get_analytics_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    month: Optional[str] = None
):
    query = {"status": {"$ne": "скасовано"}}
    if month:
        query["month"] = month
    elif start_date and end_date:
        query["order_date"] = {"$gte": start_date, "$lte": end_date}
    
    orders = await db.orders.find(query, {"_id": 0}).to_list(1000)
    
    total_revenue = sum(o.get("total_amount", 0) for o in orders)
    total_cost = sum(o.get("total_cost", 0) for o in orders)
    total_profit = sum(o.get("profit", 0) for o in orders)
    total_net_income = sum(o.get("net_income", o.get("profit", 0)) for o in orders)
    total_discount = sum(o.get("discount", 0) for o in orders)
    total_extra_income = sum(o.get("extra_income", 0) for o in orders)
    order_count = len(orders)
    avg_check = total_revenue / order_count if order_count > 0 else 0
    
    # By size
    revenue_by_size = {}
    profit_by_size = {}
    for order in orders:
        for item in order.get("items", []):
            size = item.get("size", "Інше")
            if size:
                revenue_by_size[size] = revenue_by_size.get(size, 0) + item.get("total_price", 0)
                profit_by_size[size] = profit_by_size.get(size, 0) + item.get("profit", 0)
    
    # By type
    revenue_by_type = {}
    profit_by_type = {}
    for order in orders:
        otype = order.get("order_type", "Інше")
        revenue_by_type[otype] = revenue_by_type.get(otype, 0) + order.get("total_amount", 0)
        profit_by_type[otype] = profit_by_type.get(otype, 0) + order.get("net_income", order.get("profit", 0))
    
    # By channel
    revenue_by_channel = {}
    for order in orders:
        channel = order.get("sales_channel", "Інше")
        revenue_by_channel[channel] = revenue_by_channel.get(channel, 0) + order.get("total_amount", 0)
    
    # By month
    revenue_by_month = {}
    profit_by_month = {}
    for order in orders:
        m = order.get("month", "Невідомо")
        revenue_by_month[m] = revenue_by_month.get(m, 0) + order.get("total_amount", 0)
        profit_by_month[m] = profit_by_month.get(m, 0) + order.get("net_income", order.get("profit", 0))
    
    return {
        "total_revenue": total_revenue,
        "total_cost": total_cost,
        "total_profit": total_profit,
        "total_net_income": total_net_income,
        "total_discount": total_discount,
        "total_extra_income": total_extra_income,
        "order_count": order_count,
        "avg_check": round(avg_check, 2),
        "revenue_by_size": revenue_by_size,
        "profit_by_size": profit_by_size,
        "revenue_by_type": revenue_by_type,
        "profit_by_type": profit_by_type,
        "revenue_by_channel": revenue_by_channel,
        "revenue_by_month": revenue_by_month,
        "profit_by_month": profit_by_month
    }

@api_router.get("/analytics/daily")
async def get_daily_analytics(date_str: Optional[str] = None):
    if not date_str:
        date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    query = {
        "order_date": {"$regex": f"^{date_str}"},
        "status": {"$ne": "скасовано"}
    }
    orders = await db.orders.find(query, {"_id": 0}).to_list(100)
    
    total_revenue = sum(o.get("total_amount", 0) for o in orders)
    total_net_income = sum(o.get("net_income", o.get("profit", 0)) for o in orders)
    order_count = len(orders)
    
    return {
        "date": date_str,
        "revenue": total_revenue,
        "profit": total_net_income,
        "order_count": order_count
    }

@api_router.get("/analytics/months")
async def get_available_months():
    orders = await db.orders.find({}, {"month": 1, "_id": 0}).to_list(1000)
    months = list(set(o.get("month") for o in orders if o.get("month")))
    return sorted(months, reverse=True)

# ========== EXPORT ==========

@api_router.get("/export/excel")
async def export_to_excel(
    month: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    query = {}
    if month:
        query["month"] = month
    elif start_date and end_date:
        query["order_date"] = {"$gte": start_date, "$lte": end_date}
    
    orders = await db.orders.find(query, {"_id": 0}).sort("order_date", -1).to_list(1000)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Замовлення"
    
    # Header style
    header_fill = PatternFill(start_color="C8553D", end_color="C8553D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Дата", "Місяць", "Назва картини", "Тип", "Розмір", "К-сть", 
               "Ціна продажу", "Собівартість", "Прибуток", "Канал", "Статус", "Коментар"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row_num = 2
    for order in orders:
        for item in order.get("items", [{}]):
            ws.cell(row=row_num, column=1, value=order.get("order_date", "")[:10]).border = thin_border
            ws.cell(row=row_num, column=2, value=order.get("month", "")).border = thin_border
            ws.cell(row=row_num, column=3, value=order.get("painting_name", "")).border = thin_border
            ws.cell(row=row_num, column=4, value=order.get("order_type", "")).border = thin_border
            ws.cell(row=row_num, column=5, value=item.get("size", item.get("product_name", ""))).border = thin_border
            ws.cell(row=row_num, column=6, value=item.get("quantity", 1)).border = thin_border
            ws.cell(row=row_num, column=7, value=item.get("total_price", 0)).border = thin_border
            ws.cell(row=row_num, column=8, value=item.get("total_cost", 0)).border = thin_border
            ws.cell(row=row_num, column=9, value=item.get("profit", 0)).border = thin_border
            ws.cell(row=row_num, column=10, value=order.get("sales_channel", "")).border = thin_border
            ws.cell(row=row_num, column=11, value=order.get("status", "")).border = thin_border
            ws.cell(row=row_num, column=12, value=order.get("comment", "")).border = thin_border
            row_num += 1
    
    # Adjust column widths
    for col in range(1, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"orders_{month or 'all'}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ========== MAIN ==========

@api_router.get("/")
async def root():
    return {"message": "Kuvot Art API"}

# ========== NOVA POSHTA MODELS ==========

class DimensionTemplate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    length: float  # cm
    width: float   # cm
    height: float  # cm
    weight: float  # kg

class DimensionTemplateCreate(BaseModel):
    name: str
    length: float
    width: float
    height: float
    weight: float

class NovaPoshtaSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "nova_poshta_settings"
    api_key: str
    sender_ref: Optional[str] = None
    sender_contact_ref: Optional[str] = None
    sender_address_ref: Optional[str] = None
    sender_city_ref: Optional[str] = None
    sender_phone: Optional[str] = None
    sender_name: Optional[str] = None

class TTNCreate(BaseModel):
    order_id: Optional[str] = None
    recipient_name: str
    recipient_phone: str
    recipient_city_ref: str
    recipient_city_name: Optional[str] = None
    recipient_warehouse_ref: str
    recipient_warehouse_name: Optional[str] = None
    weight: float
    length: float
    width: float
    height: float
    description: str
    cost: float
    payment_method: str = "Cash"  # Cash, NonCash
    payer_type: str = "Recipient"  # Sender, Recipient
    cod_amount: Optional[float] = None  # накладений платіж
    template_id: Optional[str] = None

class TTNResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_id: Optional[str] = None
    ttn_number: str
    ttn_ref: str
    recipient_name: str
    recipient_phone: str
    recipient_city: str
    recipient_warehouse: str
    weight: float
    description: str
    cost: float
    cod_amount: Optional[float] = None
    estimated_delivery: Optional[str] = None
    status: str = "created"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    print_url: Optional[str] = None

# ========== NOVA POSHTA API HELPERS ==========

async def nova_poshta_request(model_name: str, called_method: str, method_properties: Dict[str, Any], api_key: str = None) -> Dict:
    """Make a request to Nova Poshta API"""
    key = api_key or NOVA_POSHTA_API_KEY
    if not key:
        raise HTTPException(status_code=400, detail="API ключ Нової Пошти не налаштовано")
    
    payload = {
        "apiKey": key,
        "modelName": model_name,
        "calledMethod": called_method,
        "methodProperties": method_properties
    }
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        response = await client.post(NOVA_POSHTA_API_URL, json=payload)
        data = response.json()
        
        if not data.get("success"):
            errors = data.get("errors", ["Невідома помилка"])
            raise HTTPException(status_code=400, detail=f"Помилка Нової Пошти: {', '.join(errors)}")
        
        return data.get("data", [])

# ========== NOVA POSHTA SETTINGS ==========

@api_router.get("/nova-poshta/settings")
async def get_np_settings():
    """Get Nova Poshta settings"""
    settings = await db.nova_poshta_settings.find_one({"id": "nova_poshta_settings"}, {"_id": 0})
    if not settings:
        return {"api_key": "", "configured": False}
    # Mask API key for security
    masked_key = settings.get("api_key", "")
    if masked_key and len(masked_key) > 8:
        masked_key = masked_key[:4] + "****" + masked_key[-4:]
    return {
        **settings,
        "api_key_masked": masked_key,
        "configured": bool(settings.get("api_key"))
    }

@api_router.post("/nova-poshta/settings")
async def save_np_settings(api_key: str = None, sender_phone: str = None):
    """Save Nova Poshta API key and get sender data"""
    key = api_key or NOVA_POSHTA_API_KEY
    
    settings = {
        "id": "nova_poshta_settings",
        "api_key": key,
        "sender_phone": sender_phone
    }
    
    # Try to get counterparty (sender) data
    try:
        counterparties = await nova_poshta_request(
            "Counterparty",
            "getCounterparties",
            {"CounterpartyProperty": "Sender", "Page": "1"},
            key
        )
        
        if counterparties:
            sender = counterparties[0]
            settings["sender_ref"] = sender.get("Ref")
            settings["sender_name"] = sender.get("Description")
            settings["sender_city_ref"] = sender.get("City")
            
            # Get contact persons
            contacts = await nova_poshta_request(
                "Counterparty",
                "getCounterpartyContactPersons",
                {"Ref": sender.get("Ref"), "Page": "1"},
                key
            )
            if contacts:
                settings["sender_contact_ref"] = contacts[0].get("Ref")
            
            # Get sender addresses
            addresses = await nova_poshta_request(
                "Counterparty",
                "getCounterpartyAddresses",
                {"Ref": sender.get("Ref"), "CounterpartyProperty": "Sender"},
                key
            )
            if addresses:
                settings["sender_address_ref"] = addresses[0].get("Ref")
    except Exception as e:
        logger.error(f"Error getting sender data: {e}")
    
    await db.nova_poshta_settings.update_one(
        {"id": "nova_poshta_settings"},
        {"$set": settings},
        upsert=True
    )
    
    return {"message": "Налаштування збережено", "settings": settings}

# ========== NOVA POSHTA REFERENCE DATA ==========

@api_router.get("/nova-poshta/cities")
async def search_cities(search: str = "", limit: int = 20):
    """Search cities in Nova Poshta"""
    settings = await db.nova_poshta_settings.find_one({"id": "nova_poshta_settings"}, {"_id": 0})
    api_key = settings.get("api_key") if settings else NOVA_POSHTA_API_KEY
    
    data = await nova_poshta_request(
        "Address",
        "getCities",
        {"FindByString": search, "Limit": str(limit)}
    )
    
    return [{"ref": c.get("Ref"), "name": c.get("Description"), "area": c.get("AreaDescription")} for c in data]

@api_router.get("/nova-poshta/warehouses")
async def get_warehouses(city_ref: str, search: str = "", limit: int = 50):
    """Get warehouses (відділення) in a city"""
    settings = await db.nova_poshta_settings.find_one({"id": "nova_poshta_settings"}, {"_id": 0})
    api_key = settings.get("api_key") if settings else NOVA_POSHTA_API_KEY
    
    props = {"CityRef": city_ref, "Limit": str(limit)}
    if search:
        props["FindByString"] = search
    
    data = await nova_poshta_request("Address", "getWarehouses", props)
    
    return [{"ref": w.get("Ref"), "name": w.get("Description"), "number": w.get("Number")} for w in data]

@api_router.get("/nova-poshta/senders")
async def get_senders():
    """Get sender counterparties from user's Nova Poshta account"""
    settings = await db.nova_poshta_settings.find_one({"id": "nova_poshta_settings"}, {"_id": 0})
    if not settings or not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="API ключ не налаштовано")
    
    data = await nova_poshta_request(
        "Counterparty",
        "getCounterparties",
        {"CounterpartyProperty": "Sender", "Page": "1"},
        settings.get("api_key")
    )
    
    return [{"ref": s.get("Ref"), "name": s.get("Description"), "city": s.get("City")} for s in data]

# ========== DIMENSION TEMPLATES ==========

@api_router.get("/dimension-templates", response_model=List[DimensionTemplate])
async def get_dimension_templates():
    """Get all dimension templates"""
    templates = await db.dimension_templates.find({}, {"_id": 0}).to_list(100)
    return templates

@api_router.post("/dimension-templates", response_model=DimensionTemplate)
async def create_dimension_template(data: DimensionTemplateCreate):
    """Create a new dimension template"""
    template = DimensionTemplate(**data.model_dump())
    await db.dimension_templates.insert_one(template.model_dump())
    return template

@api_router.put("/dimension-templates/{template_id}", response_model=DimensionTemplate)
async def update_dimension_template(template_id: str, data: DimensionTemplateCreate):
    """Update a dimension template"""
    result = await db.dimension_templates.find_one_and_update(
        {"id": template_id},
        {"$set": data.model_dump()},
        return_document=True
    )
    if not result:
        raise HTTPException(status_code=404, detail="Шаблон не знайдено")
    result.pop("_id", None)
    return result

@api_router.delete("/dimension-templates/{template_id}")
async def delete_dimension_template(template_id: str):
    """Delete a dimension template"""
    result = await db.dimension_templates.delete_one({"id": template_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Шаблон не знайдено")
    return {"message": "Видалено"}

@api_router.post("/dimension-templates/seed")
async def seed_dimension_templates():
    """Seed default dimension templates for paintings"""
    existing = await db.dimension_templates.count_documents({})
    if existing > 0:
        return {"message": "Шаблони вже заповнено", "count": existing}
    
    templates = [
        {"name": "20x30 см", "length": 35, "width": 25, "height": 5, "weight": 0.5},
        {"name": "30x40 см", "length": 45, "width": 35, "height": 5, "weight": 0.8},
        {"name": "40x50 см", "length": 55, "width": 45, "height": 6, "weight": 1.2},
        {"name": "40x60 см", "length": 65, "width": 45, "height": 6, "weight": 1.5},
        {"name": "50x70 см", "length": 75, "width": 55, "height": 7, "weight": 2.0},
        {"name": "60x80 см", "length": 85, "width": 65, "height": 8, "weight": 2.8},
        {"name": "70x100 см", "length": 105, "width": 75, "height": 10, "weight": 4.0},
        {"name": "80x120 см", "length": 125, "width": 85, "height": 10, "weight": 5.5},
    ]
    
    for t in templates:
        template = DimensionTemplate(**t)
        await db.dimension_templates.insert_one(template.model_dump())
    
    return {"message": "Шаблони габаритів заповнено", "count": len(templates)}

# ========== TTN (INTERNET DOCUMENTS) ==========

@api_router.post("/nova-poshta/ttn")
async def create_ttn(data: TTNCreate):
    """Create a new TTN (shipping waybill) in Nova Poshta"""
    settings = await db.nova_poshta_settings.find_one({"id": "nova_poshta_settings"}, {"_id": 0})
    if not settings or not settings.get("api_key"):
        raise HTTPException(status_code=400, detail="API ключ Нової Пошти не налаштовано")
    
    if not settings.get("sender_ref"):
        raise HTTPException(status_code=400, detail="Дані відправника не налаштовано. Збережіть налаштування ще раз.")
    
    # Prepare TTN properties
    properties = {
        "PayerType": data.payer_type,
        "PaymentMethod": data.payment_method,
        "DateTime": datetime.now().strftime("%d.%m.%Y"),
        "CargoType": "Cargo",
        "Weight": str(data.weight),
        "ServiceType": "WarehouseWarehouse",
        "SeatsAmount": "1",
        "Description": data.description,
        "Cost": str(data.cost),
        "CitySender": settings.get("sender_city_ref"),
        "Sender": settings.get("sender_ref"),
        "SenderAddress": settings.get("sender_address_ref"),
        "ContactSender": settings.get("sender_contact_ref"),
        "SendersPhone": settings.get("sender_phone", ""),
        "CityRecipient": data.recipient_city_ref,
        "Recipient": "",  # Will be created as new
        "RecipientAddress": data.recipient_warehouse_ref,
        "ContactRecipient": "",
        "RecipientsPhone": data.recipient_phone,
        "VolumeGeneral": str(round(data.length * data.width * data.height / 1000000, 4)),
        "OptionsSeat": [
            {
                "volumetricVolume": str(round(data.length * data.width * data.height / 1000000, 4)),
                "volumetricWidth": str(data.width),
                "volumetricLength": str(data.length),
                "volumetricHeight": str(data.height),
                "weight": str(data.weight)
            }
        ]
    }
    
    # Create recipient as private person
    recipient_parts = data.recipient_name.split(" ", 2)
    recipient_props = {
        "FirstName": recipient_parts[0] if len(recipient_parts) > 0 else "",
        "MiddleName": recipient_parts[2] if len(recipient_parts) > 2 else "",
        "LastName": recipient_parts[1] if len(recipient_parts) > 1 else "",
        "Phone": data.recipient_phone,
        "CounterpartyType": "PrivatePerson",
        "CounterpartyProperty": "Recipient"
    }
    
    try:
        # Create recipient counterparty
        recipient_data = await nova_poshta_request(
            "Counterparty",
            "save",
            recipient_props,
            settings.get("api_key")
        )
        
        if recipient_data:
            properties["Recipient"] = recipient_data[0].get("Ref")
            properties["ContactRecipient"] = recipient_data[0].get("ContactPerson", {}).get("data", [{}])[0].get("Ref", "")
        
        # Add COD if specified
        if data.cod_amount and data.cod_amount > 0:
            properties["BackwardDeliveryData"] = [{
                "PayerType": "Recipient",
                "CargoType": "Money",
                "RedeliveryString": str(data.cod_amount)
            }]
        
        # Create internet document (TTN)
        ttn_data = await nova_poshta_request(
            "InternetDocument",
            "save",
            properties,
            settings.get("api_key")
        )
        
        if not ttn_data:
            raise HTTPException(status_code=400, detail="Не вдалося створити ТТН")
        
        ttn_info = ttn_data[0]
        
        # Save TTN to database
        ttn_record = {
            "id": str(uuid.uuid4()),
            "order_id": data.order_id,
            "ttn_number": ttn_info.get("IntDocNumber"),
            "ttn_ref": ttn_info.get("Ref"),
            "recipient_name": data.recipient_name,
            "recipient_phone": data.recipient_phone,
            "recipient_city": data.recipient_city_name or "",
            "recipient_warehouse": data.recipient_warehouse_name or "",
            "weight": data.weight,
            "description": data.description,
            "cost": data.cost,
            "cod_amount": data.cod_amount,
            "estimated_delivery": ttn_info.get("EstimatedDeliveryDate"),
            "status": "created",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "print_url": f"https://my.novaposhta.ua/orders/printDocument/orders[]/{ttn_info.get('Ref')}/type/pdf"
        }
        
        await db.ttns.insert_one(ttn_record)
        
        # Update order with TTN info if order_id provided
        if data.order_id:
            await db.orders.update_one(
                {"id": data.order_id},
                {"$set": {
                    "ttn_number": ttn_info.get("IntDocNumber"),
                    "ttn_ref": ttn_info.get("Ref"),
                    "delivery_status": "created"
                }}
            )
        
        return ttn_record
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error creating TTN: {e}")
        raise HTTPException(status_code=500, detail=f"Помилка створення ТТН: {str(e)}")

@api_router.get("/nova-poshta/ttns")
async def get_ttns(order_id: Optional[str] = None):
    """Get all TTNs or filter by order_id"""
    query = {}
    if order_id:
        query["order_id"] = order_id
    
    ttns = await db.ttns.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return ttns

@api_router.get("/nova-poshta/track/{ttn_number}")
async def track_ttn(ttn_number: str):
    """Track a TTN status"""
    settings = await db.nova_poshta_settings.find_one({"id": "nova_poshta_settings"}, {"_id": 0})
    api_key = settings.get("api_key") if settings else NOVA_POSHTA_API_KEY
    
    data = await nova_poshta_request(
        "TrackingDocument",
        "getStatusDocuments",
        {"Documents": [{"DocumentNumber": ttn_number}]},
        api_key
    )
    
    if data:
        status_info = data[0]
        # Update local record
        await db.ttns.update_one(
            {"ttn_number": ttn_number},
            {"$set": {"status": status_info.get("Status"), "status_code": status_info.get("StatusCode")}}
        )
        return status_info
    
    return {"error": "ТТН не знайдено"}

@api_router.get("/nova-poshta/print/{ttn_ref}")
async def get_print_url(ttn_ref: str):
    """Get print URL for a TTN"""
    return {
        "pdf_url": f"https://my.novaposhta.ua/orders/printDocument/orders[]/{ttn_ref}/type/pdf",
        "sticker_url": f"https://my.novaposhta.ua/orders/printMarkings/orders[]/{ttn_ref}/type/pdf"
    }

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
